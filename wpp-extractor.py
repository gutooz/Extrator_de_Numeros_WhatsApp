import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


INSTANCE_ID = "ID-DA-SUA-INSTANCIA"
TOKEN = "O-TOKEN-DA-ZAPI"
CLIENT_TOKEN = "TOKEN-DO-CLIENT"


base_url = f"https://api.z-api.io/instances/{INSTANCE_ID}/token/{TOKEN}"
headers = {
    "Client-Token": CLIENT_TOKEN,
    "Content-Type": "application/json"
}


def get_chats(page=1, page_size=100):
    url = f"{base_url}/chats"  
    params = {"page": page, "pageSize": page_size}
    response = requests.get(url, headers=headers, params=params)
    
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Erro {response.status_code}: {response.text}")
        return None


all_chats = []
page = 1

print("Extraindo conversas...")
while True:
    chats = get_chats(page=page)
    if not chats:
        break
    all_chats.extend(chats)
    print(f"Página {page}: {len(chats)} conversas")
    page += 1

print(f"\nTotal: {len(all_chats)} conversas")


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Conversas WhatsApp"


headers_row = ["Telefone", "Nome", "Mensagens Não Lidas", "Última Mensagem", "É Grupo", "Está Silenciado"]
ws.append(headers_row)

for cell in ws[1]:
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")


for chat in all_chats:
    
    timestamp = chat.get("lastMessageTime", "")
    if timestamp:
        try:
            data_ultima_msg = datetime.fromtimestamp(int(timestamp)).strftime("%d/%m/%Y %H:%M")
        except:
            data_ultima_msg = timestamp
    else:
        data_ultima_msg = ""
    
    ws.append([
        chat.get("phone", ""),
        chat.get("name", ""),
        chat.get("unread", "0"),
        data_ultima_msg,
        "Sim" if chat.get("isGroup", False) else "Não",
        "Sim" if chat.get("isMuted", "0") == "1" else "Não"
    ])


ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 18


filename = f"conversas_whatsapp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(filename)
print(f"\n✓ Planilha salva: {filename}")
